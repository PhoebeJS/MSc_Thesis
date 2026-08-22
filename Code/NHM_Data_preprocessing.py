import sys
sys.path.append('..')
from functions import *

import openpyxl

# Importing & inspecting data
image_dir = "data/Mammal_Osteology_held_out/Vol1"

images_sorted = sorted(f for f in os.listdir(image_dir) if f.lower().endswith(".jpg"))

# Printint them out to confirm order matches excel
for img in images_sorted:
    print(img)


# Building function to compare image with excel sheet by page number

# Here, just take what the sheet tab says & hand back a page number as a plain integer so it caan be compared against the image path later
def sheet_to_page_number(sheet_name: str) -> int:
    return int(sheet_name)


# Import and inspect the worksheet names
wb = openpyxl.load_workbook("data/Mammal_Osteology_held_out/Osteology Vol. 1 sample.xlsx")
print(wb.sheetnames)

# Inspecting filename split
filename = "NHM-UK_A_DF_ZOO_218_12_1_011_M_1.jpg"
parts = filename.split("_")
print(parts)

# Extracting page number from filename
def extract_page_number_from_filename(filename: str) -> int | None:
    parts = filename.split("_")
    
    if len(parts) <= 7:
        print(f"Warning: unexpected structure in {filename}, only {len(parts)} parts")
        return None
    
    candidate = parts[7]
    
    if not (candidate.isdigit() and len(candidate) == 3):
        print(f"Warning: part at index 7 doesn't look like a page number in {filename}: '{candidate}'")
        return None
    
    return int(candidate)

# Creating a lookup dictionary
def build_sheet_lookup(wb) -> dict:
    lookup = {}
    for sheet_name in wb.sheetnames: # For sheet names in the Wb
        page_num = sheet_to_page_number(sheet_name) # Take the sheet name & return it as an integer & assign it to the page num
        lookup[page_num] = sheet_name
    return lookup

def flatten_ws(ws) -> str:
    rows = []
    for i, row in enumerate(ws.iter_rows()): # ws.iter_rows gives one row at a team where each row is a tuple
        if i == 0:
            continue # Skipping the header row
        sliced = row[4:]   # Cutting off the first four columns
        clean_row = []
        for cell in sliced: # Transform every item in the tuple into a list of strings
            if cell.value is None:
                value = ""
            else:
                value = str(cell.value)
            clean_row.append(value)
        if all(x == "" for x in clean_row):
            continue
        rows.append(" | ".join(clean_row))
    flattened_sheets = "\n".join(rows)
    return flattened_sheets

# Resizing image
def decrease_image_size(input_path, output_path, max_dimension=3000, quality = 90):
    with Image.open(input_path) as img:
        if img.mode != "RGB":
            img = img.convert("RGB")
        img.thumbnail((max_dimension, max_dimension))
        img.save(output_path, "JPEG", quality=quality, optimize=True)

def compress_dataset_images(dataset: list[dict], output_root: str, max_dimension=3000, quality=90) -> list[dict]:
    for record in dataset:
        original_path = record["image_path"] # record is one dictionary from mammal_dataset - image_path looks up the values stored under that key
        filename = os.path.basename(original_path) # Returns the final component of a pathname
        new_path = os.path.join(output_root, record["document_id"], filename) # build new_path, preserving Vol1/Vol2/Vol3 subfolder + filename, rooted at output_root. os.path.join does this by taking multiple strings as separate arguments and gluing them together
        # python won't create missing folders when saving a file so need to save them
        folder_name = os.path.dirname(new_path)
        os.makedirs(folder_name, exist_ok=True) # os.makedirs(...) to ensure the folder exists
        decrease_image_size(original_path, new_path, max_dimension, quality)
        record["image_path"] = new_path
    return dataset


def link_images_to_sheets(image_dir, sheet_lookup, wb, volume_id):
    manifest_rows = []
    for img in sorted(os.listdir(image_dir)): # takes a folder path & returns a plain list of every file name in that folder, in alphabetical order
        if not img.lower().endswith(".jpg"): # If the file name doesn't end with .jpg, skip it (after first converting to a lower case in case it's Jpg, etc)
            continue
        page_num = extract_page_number_from_filename(img) # Using the function prev defined to take the page num from the file name
        if page_num is None: # If no page number, skip
            continue
        sheet_name = sheet_lookup.get(page_num) # Taking sheet_lookup from build_manifest, get the page number & make it the sheet name
        if sheet_name is None: # If this is not there
            print(f"No matching sheet for {img} (page {page_num})") # Then we have a problem bc no matches
            continue # Skip over
        ws = wb[sheet_name] # Select a sheet and assign it tows
        transcription = flatten_ws(ws) # Flatten the sheet & set it as the transcription
        manifest_rows.append({       # Add to the manifest of rows:
            "image_path": os.path.join(image_dir, img), # Image path taken from the image in image directory
            "transcription": transcription,
            "document_id": volume_id, # The sheet name
            "page_id": str(page_num),  # The page number
        })
    return manifest_rows


def build_full_manifest(data_root: str, volumes: list[dict]) -> list[dict]:
    all_rows = []
    for vol in volumes:
        excel_path = os.path.join(data_root, vol["excel"])
        image_dir = os.path.join(data_root, vol["folder"])
        wb = openpyxl.load_workbook(excel_path)
        sheet_lookup = build_sheet_lookup(wb) 
        manifest_rows = link_images_to_sheets(image_dir, sheet_lookup, wb, vol["folder"])
        all_rows.extend(manifest_rows)
    return all_rows

volumes = [
    {"folder": "Vol1", "excel": "Osteology Vol. 1 sample.xlsx"},
    {"folder": "Vol2", "excel": "Osteology Vol. 2 sample.xlsx"},
    {"folder": "Vol3", "excel": "Osteology Vol. 3 sample.xlsx"},
]

mammal_dataset_2 = build_full_manifest("data/Mammal_Osteology_held_out_2", volumes)
print(len(mammal_dataset_2))
print(mammal_dataset_2[0])  # sanity check before saving

with open("mammals_dataset_2.json", "w", encoding="utf-8") as f:
    json.dump(mammal_dataset_2, f, ensure_ascii=False, indent=2)

mammal_dataset_compressed_3 = compress_dataset_images(mammal_dataset_2, "data/Mammal_Osteology_held_out_compressed_3")

for record in mammal_dataset_compressed_3:
    size_mb = os.path.getsize(record["image_path"]) / (1024*1024)
    print(record["image_path"], f"{size_mb:.2f} MB")


with open("mammals_dataset_compressed_3.json", "w", encoding="utf-8") as f:
    json.dump(mammal_dataset_compressed_3, f, ensure_ascii=False, indent=2)