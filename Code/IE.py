## Information Extraction Case Study

# Using the transcriptions to perform a small species distribution model on species present within the NHM database (& potentially a subset of the BL? Still debating. Definetly NHM tho because it can then also act as a secondary test for the unstructured transcription). 


# Schema for IE

prompt_format_species = {
    "type": "array",
    "items":{
        "type":"object",
    "properties":{
        "speciesName": {
            "type": "string",
            "description": "Taxonomic species name (latin binomial) or 'spp.'",
            "default": ""
        },
        "verbatimLocality":{
            "type": "string",
            "description":"The original textual description of a location.",
            "default": ""
        },
        "date":{
            "type": "string",
            "description": "The time the species was found, spotted, collected, or identified. May take a numeric form that follows regional conventional. Example: day-month-year, month-day-year, or year-month-day. Numeric dates may use different separators, such as slashes (13/07/1895), hyphens (13-07-1895), and dots (13.07.1895). Numeric dates may also use shortened forms of the date. Example: using '95' instead of '1895' or '7' instead of '07'. Dates may also be written out in long format, where the full month is written out. Example: '13 July 1895' or 'July 13 1895'. In some cases, the day and year may also be written out. Example: 'day 13 of July 1895' or '13 July in year 1895'. Prepositions may also be used. Example: 'the 13th of July, 1895'. Only extract the year.",
            "default": ""
            },
        "verbatimDate":{
            "type": "string",
            "description": "The original textual description of date."
            }
        },
        "required": ["speciesName"],
        "additionalProperties": False
    }
}



# Load in the excel
import pandas as pd

gt_df = pd.read_excel("data/IE_species_v2.xlsx")
gt_df.head()

import openpyxl
wb = openpyxl.load_workbook("data/IE_species_v2.xlsx")
print(wb.sheetnames)


# Flattening sheet
def flatten_ws(ws) -> str:
    rows = []
    for i, row in enumerate(ws.iter_rows()): # ws.iter_rows gives one row at a team where each row is a tuple
        #if i == 0:
            #continue # Skipping the header row
        # sliced = row[3:]   # Cutting off the first two columns - maybe switch to cutting off last few columns?
        clean_row = []
        for cell in row: # Transform every item in the tuple into a list of strings
            if cell.value is None:
                value = ""
            else:
                value = str(cell.value)
            clean_row.append(value)
        if all(x == "" for x in clean_row):
            continue
        rows.append(" | ".join(clean_row))
    flattened_sheets = "\n".join(rows)
    return flattened_sheets

ie_gt_data = flatten_ws(wb['Sheet1'])

with open("ie_gt_data.json", "w", encoding="utf-8") as f:
    json.dump(ie_gt_data, f, ensure_ascii=False, indent=2)

# Grouping by document id (volume) & page id so i can pull GT rows per page for alignment later
gt_df.columns = gt_df.columns.str.strip()

gt_grouped = gt_df.groupby(["document_id", "page_id"])

########
# Extract model predicted transcriptions from the best transcriptions
# First load it in
model_transcriptions_ie = []
with open("corpus_metrics/claude_fewshot_nhm_specific_structured_run-2.jsonl", "r", encoding="utf-8") as f:
    for line in f:
        line = line.strip()
        if line:  # skip blank lines
            model_transcriptions_ie.append(json.loads(line))

# Checking keys in the dictionaries of the jsonl

all_keys = set() # Set makes a collection of unique values with no duplicates & no order
for d in model_transcriptions_ie:
    all_keys.update(d.keys())

print(all_keys)

# Need to flatten model transcription for comparability
def flatten_table(table, fallback_text="") -> str: # Take in the table & fallback_text (default = empty sting) as parameters. "-> str" is return type annotation saying this function returns a string
    # If the table parameter is empty or none and not a dictionary, return the text as is
    if not table or not isinstance(table, dict): 
        return fallback_text
    rows = table.get("rows", []) # Rows = the rows column in table. Return empty if rows not present
    flattened_rows = []  # Empty list for future values
    for row in rows:     # For a row in the gotten rows:
        clean_cells = [] # Another empty list
        for cell in row: # For a cell in the row
            clean_cells.append(cell.replace("\n", " ")) # Add to the empty list cells but repalce instances of \n with empty space
        flattened_rows.append(" | ".join(clean_cells)) # Join the cells into the flattened rows, using | to join the cells to each other (| is a separator between each item of the list)
    return "\n".join(flattened_rows) # Return the flattened rows, using \n to distinguish rows from each other (\n is a separator between each item of the list)

def extract_transcription(path, output_path=None):
  with open(path, "r", encoding="utf-8") as f:
    extracted_model_transcriptions = []
    for line in f:
      record = json.loads(line.strip())
      flat_table = flatten_table(record.get("table"))
      
      transcriptions = {
         "table": flat_table,
         "document_id": record.get("document_id"),
         "page_id": record.get("page_id")
      }

      extracted_model_transcriptions.append(transcriptions)
  if output_path:
    with open(output_path, "w", encoding="utf-8") as f:
        for transcriptions in extracted_model_transcriptions:
            f.write(json.dumps(transcriptions, ensure_ascii=False) + "\n")

  return extracted_model_transcriptions

# Changed to extract unstructured transcription too
extracted = extract_transcription(
    path = "corpus_metrics/claude_fewshot_nhm_specific_structured_run-2.jsonl",
    output_path = "extracted_transcriptions_IE.jsonl"

)



### Functions used - altered from functions/py
def build_messages_ie(user_prompt, transcription, sys_prompt=system_prompt, schema=prompt_format_species):
    messages = [
        {
            'role': 'system',
            'content': sys_prompt
        },
        {
            'role': 'user',
            'content': f"""
                {user_prompt}

                Transcription:
                {transcription}

                Provide the information following this JSON schema exactly and output JSON only: {schema}
                """
        }
    ]
    return messages

def call_openrouter_ie(user_prompt, transcription, model, temperature, sys_prompt=system_prompt, schema=prompt_format_species):
    messages = build_messages_ie(user_prompt, transcription, sys_prompt=sys_prompt, schema=schema)
    response = openrouter_client.chat.completions.create(
        model=model,
        messages=messages,
        temperature=temperature
    )
    usage = {
        "prompt_tokens": response.usage.prompt_tokens, # Extracts number of input tokens in api request
        "completion_tokens": response.usage.completion_tokens, # Extracts the number of output tokens from the LLM
        "total_tokens": response.usage.total_tokens, # Total input & output tokens
        "temperature": temperature
    }
    raw_output = response.choices[0].message.content
    return raw_output, usage



def openrouter_ie_pipeline(transcription, user_prompt, output_path, model, temperature, sys_prompt=system_prompt, schema=prompt_format_species, batch_size=10):
    # Check if output file already exists to prevent duplicates
    if os.path.exists(output_path):
        print(f"Warning: {output_path} already exists. Delete it first to avoid duplicates.")
        return
    results_buffer = []   # in-memory storage until we flush
    batch_counter = 0     # For counting how many batches we've gone through
    records = []
    with open(transcription, "r", encoding="utf-8") as f: 
        for line in f:
            record = json.loads(line.strip())
            records.append(record)
    for record in records:
        model_transcription = record["table"]
        document_id = record["document_id"]
        page_id = record["page_id"]  

    # Call openrouter api with current image and prompt
        raw_output, usage = call_openrouter_ie(
            transcription=model_transcription,
            user_prompt=user_prompt,
            sys_prompt=sys_prompt,
            schema=schema,
            model=model,
            temperature=temperature)
        # Then pass string into function
        result = string_to_json(raw_output)
        # Handle case where result is a list instead of a dictionary
        if isinstance(result, dict):
            result = [result]

        for entry in result:
            model_speciesName = entry.get("speciesName", "")
            model_verbatimLocality = entry.get("verbatimLocality", "")
            model_date = entry.get("date", "")
            model_verbatimDate = entry.get("verbatimDate", "")

            result_record = {
                "model_speciesName": model_speciesName,
                "model_verbatimLocality": model_verbatimLocality,
                "model_date": model_date,
                "model_verbatimDate": model_verbatimDate,
                "document_id": document_id,
                "page_id": page_id,
            }
            results_buffer.append(result_record)
            batch_counter += 1

        # Flush logic - checking if counter has reached batch size & should be flushed
        # NB: changing to JSONL logic since when we append json files like this with flushing, it will create an invalid JSON with two separate lists stuck together with no comma/wrapper
        if batch_counter >= batch_size:
            with open(output_path, "a", encoding="utf-8") as f:     
                # json.dump writes our records list into the file as formatted JSON
                # ensure_ascii=False: preserves special characters like & or accented letters as-is
                # indent=2: adds 2 space indentation so the JSON is human readable
                for r in results_buffer:
                    f.write(json.dumps(r, ensure_ascii=False) + "\n")
            print(f"✔ Saved batch")

            # reset buffer
            results_buffer = [] 
            batch_counter = 0
    # Final flushing block after the loop ends - catches any remaining records that didn't fill a complete batch
    if len(results_buffer) > 0:
        # Flush remaining records
        with open(output_path, "a", encoding="utf-8") as f:     
            # json.dump writes our records list into the file as formatted JSON
            # ensure_ascii=False: preserves special characters like & or accented letters as-is
            # indent=2: adds 2 space indentation so the JSON is human readable
            for record in results_buffer:
                f.write(json.dumps(record, ensure_ascii=False) + "\n")
        print(f"✔ Final batch saved")
        


#####
# Now using with openrouter
import sys
sys.path.append('..')
from functions import *

# Load in prompt
with open(Path("prompts/prompt_species_v2.txt")) as f:
    species_extraction_prompt_2 = f.read()


# Running for claude few-show, dataset specific examples, structured prompt, 0.7 temperature
openrouter_ie_pipeline(
    transcription="extracted_transcriptions_IE.jsonl",
    model="~anthropic/claude-sonnet-latest",
    output_path="model_species_extractions.jsonl",
    user_prompt=species_extraction_prompt_2,
    temperature=0.7
)


model_df = pd.read_json("model_species_extractions.jsonl", lines=True) # set to lines=true because regular read.json cannot handle jsonl (which is technically one json object per line)

model_df = model_df.rename(columns=
                           {"model_speciesName": "speciesName",
                             "model_verbatimLocality": "verbatimLocality",
                             "model_date": "date",
                             "model_verbatimDate": "verbatimDate"})

model_grouped = model_df.groupby(["document_id", "page_id"])

def per_page_comparison(gt_data, model_data, key): # takes in the GT & model data plus the 'key' - i.e., what variable of focus for calculations

    # Pull out the subset of rows that belong to one page (that match by document id & page id_ from the grouped data)
    gt_group = gt_data.get_group(key)       
    model_group = model_data.get_group(key) 

    # take the key (dates, species, etc) from gt data & convert it to a list since Counter needs an ordinary iterable not a pandas object
    # Change the value depending on the entity being extracted
    gt_species = gt_group["verbatimDate"].tolist()        
    model_species = model_group["verbatimDate"].tolist()

    #gt_species = gt_group["date"].astype(str).tolist()
    #model_species = model_group["date"].astype(str).tolist()
    
    # Turns the list into counts per unique value (eg: ["a", "a, "b]) into {"a": 2. "b": 1}
    gt_counter = Counter(gt_species)       
    model_counter = Counter(model_species)

    # Finds the intersection for counter. For each value present in both, it keeps the smaller of the two counts. This makes what repeats count correctly (eg: 3 GT "Species A" and 2 predicted "Species A" → 2 correct, not 3)
    correct = model_counter & gt_counter

    # Sum the total counts in each counter
    return {
        "n_correct": sum(correct.values()),  # Total number of correct values (i.e., where intersections/matches were found)
        "n_predicted": sum(model_counter.values()), # The number of predicted unique values
        "n_actual": sum(gt_counter.values()) # The number of actual unique values
    }


results = []
for key in gt_grouped.groups.keys(): # in the GT group, look at the groups & then the keys in the group
    if key in model_grouped.groups: # If the key is present in the groups
        results.append(per_page_comparison(gt_grouped, model_grouped, key)) # Apply the per page comparison to find the number of unique values & add to the results
    else:
        # page had GT species but no model predictions at all
        gt_count = len(gt_grouped.get_group(key))
        results.append({"n_correct": 0, "n_predicted": 0, "n_actual": gt_count})


def bootstrap_ci(scores, n_bootstrap=2000, ci_level=95):
    if len(scores) < 2:
        return (float("nan"), float("nan"))
    scores = np.array(scores)
    bootstrap_means = []
    for _ in range(n_bootstrap):
        resample = np.random.choice(scores, size=len(scores), replace=True)
        bootstrap_means.append(np.mean(resample))
    lower_p = (100 - ci_level) / 2
    upper_p = 100 - lower_p
    ci_lower = np.percentile(bootstrap_means, lower_p)
    ci_upper = np.percentile(bootstrap_means, upper_p)
    return (ci_lower, ci_upper)

total_correct = sum(r["n_correct"] for r in results)
total_predicted = sum(r["n_predicted"] for r in results)
total_actual = sum(r["n_actual"] for r in results)

precision = total_correct / total_predicted if total_predicted > 0 else 0 # True positives divided by false positives + true positives
recall = total_correct / total_actual if total_actual > 0 else 0 # True positives divided by true positives and false negatives
f1 = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0

print(f"Precision: {precision:.3f}, Recall: {recall:.3f}, F1: {f1:.3f}")



# Recall CI
per_page_recall = [
    r["n_correct"] / r["n_actual"] if r["n_actual"] > 0 else float("nan")
    for r in results
]
per_page_recall_clean = [x for x in per_page_recall if not math.isnan(x)]
recall_ci_lower, recall_ci_upper = bootstrap_ci(per_page_recall_clean)

# Precision CI
per_page_precision = [
    r["n_correct"] / r["n_predicted"] if r["n_predicted"] > 0 else float("nan")
    for r in results
]
per_page_precision_clean = [x for x in per_page_precision if not math.isnan(x)]
precision_ci_lower, precision_ci_upper = bootstrap_ci(per_page_precision_clean)


# F1 CI
per_page_f1 = []
for r in results:
    precision = r["n_correct"] / r["n_predicted"] if r["n_predicted"] > 0 else float("nan")
    recall = r["n_correct"] / r["n_actual"] if r["n_actual"] > 0 else float("nan")
    f1 = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else float("nan")
    per_page_f1.append(f1)

per_page_f1_clean = [x for x in per_page_f1 if not math.isnan(x)]
f1_ci_lower, f1_ci_upper = bootstrap_ci(per_page_f1_clean)


print(f"Precision: {precision:.3f} (95% CI: {precision_ci_lower:.3f}–{precision_ci_upper:.3f})")
print(f"Recall: {recall:.3f} (95% CI: {recall_ci_lower:.3f}–{recall_ci_upper:.3f})")
print(f"F1: {f1:.3f} (95% CI: {f1_ci_lower:.3f}–{f1_ci_upper:.3f})")



####### Then same thing for unstructured transcription
def extract_transcription_2(path, output_path=None):
  with open(path, "r", encoding="utf-8") as f:
    extracted_model_transcriptions = []
    for line in f:
      record = json.loads(line.strip())
      flat_table = flatten_table(record.get("table"))
      
      transcriptions = {
        "model_transcription": record.get("model_transcription"),
         "table": flat_table,
         "document_id": record.get("document_id"),
         "page_id": record.get("page_id")
      }

      extracted_model_transcriptions.append(transcriptions)
  if output_path:
    with open(output_path, "w", encoding="utf-8") as f:
        for transcriptions in extracted_model_transcriptions:
            f.write(json.dumps(transcriptions, ensure_ascii=False) + "\n")

  return extracted_model_transcriptions

# Extract the model transcriptions into usable form 
extracted = extract_transcription_2(
    path = "corpus_metrics/claude_fewshot_nhm_unstructured_run-1.jsonl",
    output_path = "unstructured_extracted_transcriptions_IE.jsonl"

)

def openrouter_ie_pipeline_2(transcription, user_prompt, output_path, model, temperature, sys_prompt=system_prompt, schema=prompt_format_species, batch_size=10):
    # Check if output file already exists to prevent duplicates
    if os.path.exists(output_path):
        print(f"Warning: {output_path} already exists. Delete it first to avoid duplicates.")
        return
    results_buffer = []   # in-memory storage until we flush
    batch_counter = 0     # For counting how many batches we've gone through
    records = []
    with open(transcription, "r", encoding="utf-8") as f: 
        for line in f:
            record = json.loads(line.strip())
            records.append(record)
    for record in records:
        model_transcription = record["model_transcription"]
        document_id = record["document_id"]
        page_id = record["page_id"]  

    # Call openrouter api with current image and prompt
        raw_output, usage = call_openrouter_ie(
            transcription=model_transcription,
            user_prompt=user_prompt,
            sys_prompt=sys_prompt,
            schema=schema,
            model=model,
            temperature=temperature)
        # Then pass string into function
        result = string_to_json(raw_output)
        # Handle case where result is a list instead of a dictionary
        if isinstance(result, dict):
            result = [result]

        for entry in result:
            model_speciesName = entry.get("speciesName", "")
            model_verbatimLocality = entry.get("verbatimLocality", "")
            model_date = entry.get("date", "")
            model_verbatimDate = entry.get("verbatimDate", "")

            result_record = {
                "model_speciesName": model_speciesName,
                "model_verbatimLocality": model_verbatimLocality,
                "model_date": model_date,
                "model_verbatimDate": model_verbatimDate,
                "document_id": document_id,
                "page_id": page_id,
            }
            results_buffer.append(result_record)
            batch_counter += 1

        # Flush logic - checking if counter has reached batch size & should be flushed
        # NB: changing to JSONL logic since when we append json files like this with flushing, it will create an invalid JSON with two separate lists stuck together with no comma/wrapper
        if batch_counter >= batch_size:
            with open(output_path, "a", encoding="utf-8") as f:     
                # json.dump writes our records list into the file as formatted JSON
                # ensure_ascii=False: preserves special characters like & or accented letters as-is
                # indent=2: adds 2 space indentation so the JSON is human readable
                for r in results_buffer:
                    f.write(json.dumps(r, ensure_ascii=False) + "\n")
            print(f"✔ Saved batch")

            # reset buffer
            results_buffer = [] 
            batch_counter = 0
    # Final flushing block after the loop ends - catches any remaining records that didn't fill a complete batch
    if len(results_buffer) > 0:
        # Flush remaining records
        with open(output_path, "a", encoding="utf-8") as f:     
            # json.dump writes our records list into the file as formatted JSON
            # ensure_ascii=False: preserves special characters like & or accented letters as-is
            # indent=2: adds 2 space indentation so the JSON is human readable
            for record in results_buffer:
                f.write(json.dumps(record, ensure_ascii=False) + "\n")
        print(f"✔ Final batch saved")
        
openrouter_ie_pipeline_2(
    transcription="unstructured_extracted_transcriptions_IE.jsonl",
    model="~anthropic/claude-sonnet-latest",
    output_path="unstructured_model_species_extractions.jsonl",
    user_prompt=species_extraction_prompt_2,
    temperature=0.7
)

model_df_2 = pd.read_json("unstructured_model_species_extractions.jsonl", lines=True) # set to lines=true because regular read.json cannot handle jsonl (which is technically one json object per line)

model_df_2 = model_df_2.rename(columns=
                           {"model_speciesName": "speciesName",
                             "model_verbatimLocality": "verbatimLocality",
                             "model_date": "date",
                             "model_verbatimDate": "verbatimDate"})

model_grouped_2 = model_df_2.groupby(["document_id", "page_id"])

def per_page_comparison(gt_data, model_data, key): # takes in the GT & model data plus the 'key' - i.e., what variable of focus for calculations

    # Pull out the subset of rows that belong to one page (that match by document id & page id_ from the grouped data)
    gt_group = gt_data.get_group(key)       
    model_group = model_data.get_group(key) 

    # take the key (dates, species, etc) from gt data & convert it to a list since Counter needs an ordinary iterable not a pandas object
    gt_species = gt_group["verbatimDate"].tolist()        
    model_species = model_group["verbatimDate"].tolist()

    #gt_species = gt_group["date"].astype(str).tolist()
    #model_species = model_group["date"].astype(str).tolist()
    
    # Turns the list into counts per unique value (eg: ["a", "a, "b]) into {"a": 2. "b": 1}
    gt_counter = Counter(gt_species)       
    model_counter = Counter(model_species)

    # Finds the intersection for counter. For each value present in both, it keeps the smaller of the two counts. This makes what repeats count correctly (eg: 3 GT "Species A" and 2 predicted "Species A" → 2 correct, not 3)
    correct = model_counter & gt_counter

    # Sum the total counts in each counter
    return {
        "n_correct": sum(correct.values()),  # Total number of correct values (i.e., where intersections/matches were found)
        "n_predicted": sum(model_counter.values()), # The number of predicted unique values
        "n_actual": sum(gt_counter.values()) # The number of actual unique values
    }

results = []
for key in gt_grouped.groups.keys(): # in the GT group, look at the groups & then the keys in the group
    if key in model_grouped_2.groups: # If the key is present in the groups
        results.append(per_page_comparison(gt_grouped, model_grouped_2, key)) # Apply the per page comparison to find the number of unique values & add to the results
    else:
        # page had GT species but no model predictions at all
        gt_count = len(gt_grouped.get_group(key))
        results.append({"n_correct": 0, "n_predicted": 0, "n_actual": gt_count})

total_correct = sum(r["n_correct"] for r in results)
total_predicted = sum(r["n_predicted"] for r in results)
total_actual = sum(r["n_actual"] for r in results)

precision = total_correct / total_predicted if total_predicted > 0 else 0 # True positives divided by false positives + true positives
recall = total_correct / total_actual if total_actual > 0 else 0 # True positives divided by true positives and false negatives
f1 = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0


print(f"Precision: {precision:.3f}, Recall: {recall:.3f}, F1: {f1:.3f}")


# Recall CI
per_page_recall = [
    r["n_correct"] / r["n_actual"] if r["n_actual"] > 0 else float("nan")
    for r in results
]
per_page_recall_clean = [x for x in per_page_recall if not math.isnan(x)]
recall_ci_lower, recall_ci_upper = bootstrap_ci(per_page_recall_clean)

# Precision CI
per_page_precision = [
    r["n_correct"] / r["n_predicted"] if r["n_predicted"] > 0 else float("nan")
    for r in results
]
per_page_precision_clean = [x for x in per_page_precision if not math.isnan(x)]
precision_ci_lower, precision_ci_upper = bootstrap_ci(per_page_precision_clean)


# F1 CI
per_page_f1 = []
for r in results:
    precision = r["n_correct"] / r["n_predicted"] if r["n_predicted"] > 0 else float("nan")
    recall = r["n_correct"] / r["n_actual"] if r["n_actual"] > 0 else float("nan")
    f1 = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else float("nan")
    per_page_f1.append(f1)

per_page_f1_clean = [x for x in per_page_f1 if not math.isnan(x)]
f1_ci_lower, f1_ci_upper = bootstrap_ci(per_page_f1_clean)


print(f"Precision: {precision:.3f} (95% CI: {precision_ci_lower:.3f}–{precision_ci_upper:.3f})")
print(f"Recall: {recall:.3f} (95% CI: {recall_ci_lower:.3f}–{recall_ci_upper:.3f})")
print(f"F1: {f1:.3f} (95% CI: {f1_ci_lower:.3f}–{f1_ci_upper:.3f})")